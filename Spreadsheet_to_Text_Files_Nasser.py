import openpyxl
import pyinputplus as pyip

wb = openpyxl.load_workbook('Ditto.xlsx')
sheet = wb.active

print('I am creating ' + str(sheet.max_column) + ' text files.')

for l in range(sheet.max_column):
    file_location = pyip.inputStr(prompt=r'What text file directory will you use? eg. C:\\Users\\medis\\dfe.txt: ')
    lines = []
    with open(file_location, 'w') as fr:                # Loop that reads the text file lines
        for i in range(len(list(sheet.columns)[l])):                     # Loop that writes to a spreadsheet
            lines.append(sheet.cell(row=i+1, column=l+1).value)
        fr.write("\n".join(str(n) for n in lines) + "\n")

        print(lines)

