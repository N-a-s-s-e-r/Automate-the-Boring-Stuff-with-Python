import openpyxl, sys

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb['Sheet2']

if len(sys.argv) > 1:
    n = int(sys.argv[1])
    m = n+2

    for Num1 in range(1, m):    # skip the first row
        sheet.cell(row=Num1, column=1).value = sheet.cell(row=Num1, column=Num2).value
        sheet.cell(row=1, column=Num1).value = sheet.cell(row=Num1, column=Num2).value

        for Num2 in range(1, sheet.max_column):
            sheet.cell(row=Num1, column=Num2).value = sheet.cell(row=Num2, column=Num1).value

    wb.save('updatedexample.xlsx')
    print('I\'m done, Check your working directory to see the magic of this program')

else:
    print('Provide a number with your program')
