import openpyxl, sys

wb = openpyxl.Workbook() # Create a blank workbook.

sheet = wb.active # Get the active sheet.

if len(sys.argv) > 1:
    m = int(sys.argv[1]) + 1

    for Num1 in range(1, m):
        sheet.cell(row=Num1, column=1).value = Num1 # Skip the first column.
        sheet.cell(row=1, column=Num1).value = Num1 # Skip the first row.

        for Num2 in range(1, sheet.max_column):
            sheet.cell(row=Num1, column=Num2).value = Num1 * Num2
            sheet.cell(row=Num1, column=Num1).value = Num1 * Num1
            sheet.cell(row=Num2, column=Num1).value = Num1 * Num2

    wb.save(f'Multiplication_Table_{m-1}.xlsx')
    print('I\'m done, Check your working directory to see the magic of this program')

else:
    print('Provide a number with your program')
