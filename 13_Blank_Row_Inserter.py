import openpyxl, sys

File_Name = str(sys.argv[3])
wb = openpyxl.load_workbook(File_Name)

sheet = wb['Sheet2']
sheet1 = wb['Sheet3']

if len(sys.argv) > 1:
    a = int(sys.argv[1])
    n = a-1
    b = int(sys.argv[2])
    m = b+n+1
    p = m+1

    for Num1 in range(1, n+1):#if 4, 1to3
        for Num2 in range(1, sheet.max_column):
            sheet1.cell(row=Num1, column=Num2).value = sheet.cell(row=Num1, column=Num2).value
    for Num1 in range(a, m+1):#if 3, 4,5,6,7
        for Num2 in range(1, sheet.max_column):
            sheet1.cell(row=Num1, column=Num2).value = None
    for Num1 in range(a, sheet.max_row+1):
        for Num2 in range(1, sheet.max_column):
            sheet1.cell(row=Num1+b, column=Num2).value = sheet.cell(row=Num1, column=Num2).value

    wb.save('updatedexample1.xlsx')
    print('I\'m done, Check your working directory to see the magic of this program')

else:
    print('Provide a number with your program')
# Write your code here :-)
